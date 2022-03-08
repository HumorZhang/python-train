import json

import openpyxl
import pymysql

def writeExcel():
    outwb = openpyxl.Workbook()  # 打开一个将写的文件

    outws = outwb.create_sheet(index=0)
    i = 1  # 注意：'cell'函数中行列起始值为1

    outws.cell(column = 1 , row = 1 , value = "123123123")

    savexlsx = "results.xls"
    outwb.save(savexlsx)
def parse():
    with open("station.txt", "r", encoding='utf-8') as f:
        data = json.loads(f.read())    # load的传入参数为字符串类型
        for i  in range(len(data['city'])):
            for j in range(len(data['city'][i]['list'])):
                print(data['city'][i]['list'][j]['name'])


def get_arg():
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)
    #指定file以utf-8的格式打开

    k=1
    with open("formation", "r", encoding='utf-8') as f:
        list = f.readlines()
        for i in range(0, len(list)):
            list[i] = list[i].rstrip('\n')
        for i in range(len(list)):
            for j in range(len(list)):
                if(i==j):
                    continue
                else:
                    print(list[i],list[j])
                    outws.cell(column = k , row = 1 , value = list[i])
                    outws.cell(column = k , row = 2 , value = list[j])
                    outws.cell(column = k , row = 3 , value = "0")
                    k=k+1
    print(k)
    f.close()
    savexlsx = "results.xls"
    outwb.save(savexlsx)


if __name__ == '__main__':
    # parse()
     get_arg()
    # writeExcel()
